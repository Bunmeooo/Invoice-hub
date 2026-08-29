import os
import glob
import sqlite3
import datetime
import json
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import List, Dict, Any

RECIPIENT_EMAIL = "hznguyen1993@gmail.com"
REPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "weekly_feedback_reports")

def get_all_cross_user_feedback(days: int = 7) -> List[Dict[str, Any]]:
    """Collects feedback from all user databases across the entire workspace."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    db_paths = glob.glob(os.path.join(base_dir, "user_data", "*", "invoices.db"))
    default_db = os.path.join(base_dir, "invoices.db")
    if os.path.exists(default_db):
        db_paths.append(default_db)
        
    all_feedbacks = []
    
    for db_file in db_paths:
        try:
            conn = sqlite3.connect(db_file)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            # Ensure status and admin_note columns exist
            cursor.execute("PRAGMA table_info(user_feedback)")
            cols = [c[1] for c in cursor.fetchall()]
            if "status" not in cols:
                cursor.execute("ALTER TABLE user_feedback ADD COLUMN status TEXT DEFAULT 'Chờ xử lý'")
            if "admin_note" not in cols:
                cursor.execute("ALTER TABLE user_feedback ADD COLUMN admin_note TEXT DEFAULT ''")
            conn.commit()
            
            cursor.execute("""
                SELECT id, user_id, rating, category, comment, created_at, status, admin_note
                FROM user_feedback
                ORDER BY id DESC
            """)
            rows = cursor.fetchall()
            for r in rows:
                item = dict(r)
                item["db_path"] = db_file
                if not item.get("status"):
                    item["status"] = "Chờ xử lý"
                if not item.get("admin_note"):
                    item["admin_note"] = ""
                all_feedbacks.append(item)
            conn.close()
        except Exception as e:
            continue
            
    return all_feedbacks

def delete_cross_user_feedback(db_path: str, feedback_id: int) -> bool:
    """Xóa bình luận cụ thể theo file DB và feedback_id."""
    try:
        if not os.path.exists(db_path):
            return False
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("DELETE FROM user_feedback WHERE id = ?", (feedback_id,))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Error deleting feedback {feedback_id} from {db_path}: {e}")
        return False

def update_cross_user_feedback(db_path: str, feedback_id: int, status: str, admin_note: str = "") -> bool:
    """Cập nhật trạng thái và ghi chú xử lý của Quản trị viên."""
    try:
        if not os.path.exists(db_path):
            return False
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("""
            UPDATE user_feedback 
            SET status = ?, admin_note = ? 
            WHERE id = ?
        """, (status, admin_note, feedback_id))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Error updating feedback {feedback_id} in {db_path}: {e}")
        return False

def generate_weekly_report(days: int = 7) -> Dict[str, Any]:
    """Generates summary statistics and formatted content for the weekly feedback report."""
    feedbacks = get_all_cross_user_feedback(days=days)
    now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    start_date = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime("%d/%m/%Y")
    
    total_cnt = len(feedbacks)
    avg_rating = round(sum(f.get("rating", 5) for f in feedbacks) / total_cnt, 2) if total_cnt > 0 else 5.0
    
    cat_counts = {}
    for f in feedbacks:
        cat = f.get("category", "💡 Góp ý khác")
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
        
    # Build Markdown Content
    md = f"# 📊 BÁO CÁO TỔNG HỢP GÓP Ý & PHẢN HỒI NGƯỜI DÙNG HÀNG TUẦN\n"
    md += f"**Hệ Thống:** INVOICE HUB | FINTECH ENTERPRISE\n"
    md += f"**Gửi tới:** `{RECIPIENT_EMAIL}`\n"
    md += f"**Thời gian báo cáo:** {start_date} - {now_str}\n\n"
    md += f"### 📈 1. TỔNG QUAN CHỈ SỐ HÀI LÒNG:\n"
    md += f"- **Tổng số lượt góp ý trong tuần:** `{total_cnt}` lượt\n"
    md += f"- **Điểm đánh giá trung bình:** `{avg_rating} / 5.0 ⭐`\n"
    md += f"- **Phân loại ý kiến đóng góp:**\n"
    if cat_counts:
        for cat, cnt in cat_counts.items():
            md += f"  • **{cat}:** `{cnt}` lượt ({round(cnt/total_cnt*100, 1)}%)\n"
    else:
        md += f"  • _Chưa có dữ liệu phản hồi mới trong tuần này._\n"
        
    md += f"\n### 📝 2. CHI TIẾT Ý KIẾN ĐÓNG GÓP TỪ TỪNG NGƯỜI DÙNG:\n\n"
    if not feedbacks:
        md += f"> [!NOTE]\n> Hiện tại không có phản hồi lỗi hoặc đề xuất tính năng mới nào trong tuần qua. Toàn bộ hệ thống vận hành ổn định 100%.\n"
    else:
        from content_moderator import mask_profanity
        md += "| STT | Người Dùng | Đánh Giá | Phân Loại | Nội Dung Góp Ý / Yêu Cầu Cải Tiến | Thời Gian |\n"
        md += "| :--- | :--- | :--- | :--- | :--- | :--- | \n"
        for i, f in enumerate(feedbacks, 1):
            stars = "⭐" * f.get("rating", 5)
            user = f.get("user_id", "Kế toán")
            cat = f.get("category", "-")
            cmt = mask_profanity(f.get("comment", "").replace("\n", " ").strip())
            time_val = str(f.get("created_at") or "-")
            md += f"| {i} | `{user}` | {stars} | **{cat}** | {cmt} | `{time_val}` |\n"
            
    # Build HTML Content for Email
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; line-height: 1.6; color: #1E293B; background: #F8FAFC; padding: 20px; }}
            .card {{ background: #FFFFFF; border: 1px solid #E2E8F0; border-radius: 12px; padding: 24px; max-width: 800px; margin: 0 auto; box-shadow: 0 4px 12px rgba(0,0,0,0.05); }}
            .header {{ border-bottom: 2px solid #3B82F6; padding-bottom: 16px; margin-bottom: 20px; }}
            .title {{ font-size: 20px; font-weight: 800; color: #1E3A8A; margin: 0; }}
            .subtitle {{ font-size: 13px; color: #64748B; margin-top: 4px; }}
            .kpi-grid {{ display: flex; gap: 16px; margin: 20px 0; }}
            .kpi-box {{ flex: 1; background: #EFF6FF; border: 1px solid #BFDBFE; border-radius: 8px; padding: 14px; text-align: center; }}
            .kpi-val {{ font-size: 22px; font-weight: 800; color: #1D4ED8; }}
            .kpi-lbl {{ font-size: 12px; color: #475569; font-weight: 600; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 16px; font-size: 13px; }}
            th, td {{ padding: 10px 12px; border: 1px solid #E2E8F0; text-align: left; }}
            th {{ background: #F1F5F9; color: #334155; font-weight: 700; }}
            tr:nth-child(even) {{ background: #F8FAFC; }}
            .badge {{ display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 700; background: #DBEAFE; color: #1E40AF; }}
            .footer {{ margin-top: 24px; padding-top: 16px; border-top: 1px solid #E2E8F0; font-size: 12px; color: #94A3B8; text-align: center; }}
        </style>
    </head>
    <body>
        <div class="card">
            <div class="header">
                <div class="title">📊 BÁO CÁO TỔNG HỢP Ý KIẾN NGƯỜI DÙNG HÀNG TUẦN</div>
                <div class="subtitle">Hệ thống Quản lý & Bóc tách Hóa đơn Điện tử (INVOICE HUB) • {start_date} - {now_str}</div>
            </div>
            
            <div class="kpi-grid">
                <div class="kpi-box">
                    <div class="kpi-val">{total_cnt}</div>
                    <div class="kpi-lbl">Tổng Ý Kiến Góp Ý</div>
                </div>
                <div class="kpi-box">
                    <div class="kpi-val">{avg_rating} ⭐</div>
                    <div class="kpi-lbl">Mức Độ Hài Lòng</div>
                </div>
                <div class="kpi-box">
                    <div class="kpi-val">{len(cat_counts)}</div>
                    <div class="kpi-lbl">Nhóm Chủ Đề</div>
                </div>
            </div>
            
            <h4 style="margin-top: 20px; color: #1E3A8A;">📝 Danh Sách Chi Tiết Ý Kiến Cần Cải Tiến / Nâng Cấp:</h4>
            <table>
                <thead>
                    <tr>
                        <th>STT</th>
                        <th>Người Dùng</th>
                        <th>Đánh Giá</th>
                        <th>Phân Loại</th>
                        <th>Nội Dung Chi Tiết</th>
                        <th>Thời Gian</th>
                    </tr>
                </thead>
                <tbody>
    """
    if not feedbacks:
        html += """
                    <tr>
                        <td colspan="6" style="text-align: center; color: #64748B; padding: 20px;">
                            🎉 Chưa có phản hồi lỗi mới trong tuần qua. Toàn bộ tính năng đang hoạt động hoàn hảo!
                        </td>
                    </tr>
        """
    else:
        for i, f in enumerate(feedbacks, 1):
            stars = "⭐" * f.get("rating", 5)
            html += f"""
                    <tr>
                        <td>{i}</td>
                        <td><b>{f.get('user_id', 'Kế toán')}</b></td>
                        <td>{stars}</td>
                        <td><span class="badge">{f.get('category', '-')}</span></td>
                        <td>{f.get('comment', '')}</td>
                        <td style="color: #64748B; font-size: 11px;">{f.get('created_at', '-')}</td>
                    </tr>
            """
            
    html += f"""
                </tbody>
            </table>
            
            <div class="footer">
                Báo cáo tự động tổng hợp định kỳ gửi tới <b>{RECIPIENT_EMAIL}</b>.<br>
                INVOICE HUB | FINTECH ENTERPRISE © {datetime.datetime.now().year}
            </div>
        </div>
    </body>
    </html>
    """
    
    return {
        "total_count": total_cnt,
        "avg_rating": avg_rating,
        "category_counts": cat_counts,
        "feedbacks": feedbacks,
        "markdown": md,
        "html": html,
        "timestamp": now_str
    }

def save_weekly_report_snapshot():
    """Saves report to local markdown and JSON files for historical logging."""
    os.makedirs(REPORT_DIR, exist_ok=True)
    report_data = generate_weekly_report(days=7)
    date_tag = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    
    md_file = os.path.join(REPORT_DIR, f"weekly_feedback_report_{date_tag}.md")
    with open(md_file, "w", encoding="utf-8") as f:
        f.write(report_data["markdown"])
        
    latest_file = os.path.join(REPORT_DIR, "latest_digest.json")
    with open(latest_file, "w", encoding="utf-8") as f:
        json.dump({
            "total_count": report_data["total_count"],
            "avg_rating": report_data["avg_rating"],
            "category_counts": report_data["category_counts"],
            "timestamp": report_data["timestamp"],
            "feedbacks": report_data["feedbacks"]
        }, f, ensure_ascii=False, indent=2)
        
    return md_file

def generate_admin_analytics_excel() -> bytes:
    """
    Xuat file Excel tong hop da chieu danh cho Quan tri vien:
    Sheet 1: Danh Sach Nguoi Dung (User Directory)
    Sheet 2: Thong Ke Su Dung Tinh Nang (Feature Usage & Workload)
    Sheet 3: Gop Y & Bao Loi (User Feedback & Bug Tracking)
    Sheet 4: Tong Quan Phan Tich (Executive Insights & Recommendations)
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from auth import get_all_registered_users
    
    users = get_all_registered_users()
    feedbacks = get_all_cross_user_feedback(days=365)
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Dinh dang Style chuan Doanh nghiep
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Times New Roman", size=14, bold=True, color="1E3A8A")
    subtitle_font = Font(name="Times New Roman", size=10, italic=True, color="475569")
    regular_font = Font(name="Times New Roman", size=11)
    bold_font = Font(name="Times New Roman", size=11, bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # -------------------------------------------------------------
    # SHEET 1: DANH SÁCH TÀI KHOẢN NGƯỜI DÙNG
    # -------------------------------------------------------------
    ws1 = wb.create_sheet(title="1. Danh Sách Người Dùng")
    ws1.views.sheetView[0].showGridLines = True
    
    ws1["A1"] = "BẢNG TỔNG HỢP DANH SÁCH TÀI KHOẢN ĐĂNG KÝ HỆ THỐNG"
    ws1["A1"].font = title_font
    ws1["A2"] = f"Thời gian trích xuất: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | INVOICE HUB"
    ws1["A2"].font = subtitle_font
    
    headers1 = [
        "STT", "Tên Đăng Nhập (Username)", "Họ và Tên / Đơn Vị", "Số Điện Thoại / Zalo",
        "Email Liên Hệ", "Tên Doanh Nghiệp / Công Ty", "Ngày Đăng Ký (Bắt Đầu)",
        "Hạn Kết Thúc (30 Ngày)", "Số Ngày Còn Lại", "Đăng Nhập Gần Nhất", "Gói Dịch Vụ", "Trạng Thái"
    ]
    
    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=4, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    for row_idx, u in enumerate(users, 5):
        d_left = u.get("days_remaining", 0)
        is_adm = (u.get("username") == "hznguyen1997")
        d_str = "Vĩnh viễn (Admin)" if is_adm else f"{d_left} ngày"
        stt_status = "Hoạt động" if (d_left > 0 or is_adm) else "Hết hạn dùng thử"
        
        row_vals = [
            row_idx - 4,
            u.get("username", ""),
            u.get("full_name", ""),
            u.get("phone", ""),
            u.get("email", ""),
            u.get("company", ""),
            u.get("registered_at", ""),
            u.get("trial_end", ""),
            d_str,
            u.get("last_login", ""),
            "👑 Admin" if is_adm else "🎁 Dùng thử 30 ngày (Full)",
            stt_status
        ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx in [1, 7, 8, 9, 10, 11, 12]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    # -------------------------------------------------------------
    # SHEET 2: THỐNG KÊ SỬ DỤNG TÍNH NĂNG
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="2. Thống Kê Sử Dụng")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2["A1"] = "BÁO CÁO THỐNG KÊ TẦN SUẤT & KHỐI LƯỢNG SỬ DỤNG TÍNH NĂNG"
    ws2["A1"].font = title_font
    ws2["A2"] = "Dữ liệu đo lường khối lượng bóc tách hóa đơn, doanh số thẩm định và tương tác của từng người dùng"
    ws2["A2"].font = subtitle_font
    
    headers2 = [
        "STT", "Tên Tài Khoản", "Họ và Tên", "Số HĐ Đã Bóc Tách", "Số NCC Đã Giao Dịch",
        "Tổng Tiền Chưa Thuế (VND)", "Tổng Thuế GTGT (VND)", "Tổng Thanh Toán Thẩm Định (VND)", "Mức Độ Hoạt Động"
    ]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    for row_idx, u in enumerate(users, 5):
        uname = u.get("username", "")
        u_db = os.path.join(base_dir, "user_data", uname, "invoices.db")
        inv_cnt, ncc_cnt, pre_tax, vat_sum, tot_sum = 0, 0, 0, 0, 0
        if os.path.exists(u_db):
            try:
                conn_u = sqlite3.connect(u_db)
                c_u = conn_u.cursor()
                c_u.execute("SELECT count(*), count(distinct mst_ban), sum(tong_tien_chua_thue), sum(tien_thue), sum(tong_tien) FROM invoices")
                row_db = c_u.fetchone()
                if row_db:
                    inv_cnt = row_db[0] or 0
                    ncc_cnt = row_db[1] or 0
                    pre_tax = row_db[2] or 0
                    vat_sum = row_db[3] or 0
                    tot_sum = row_db[4] or 0
                conn_u.close()
            except Exception:
                pass
                
        activity_level = "Tích cực" if inv_cnt >= 10 else ("Đang trải nghiệm" if inv_cnt > 0 else "Mới khởi tạo")
        row_vals2 = [
            row_idx - 4,
            uname,
            u.get("full_name", ""),
            inv_cnt,
            ncc_cnt,
            pre_tax,
            vat_sum,
            tot_sum,
            activity_level
        ]
        for col_idx, val in enumerate(row_vals2, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx in [1, 9]:
                cell.alignment = center_align
            elif col_idx in [4, 5]:
                cell.alignment = right_align
                cell.number_format = "#,##0"
            elif col_idx in [6, 7, 8]:
                cell.alignment = right_align
                cell.number_format = "#,##0"
            else:
                cell.alignment = left_align
                
    # -------------------------------------------------------------
    # SHEET 3: GÓP Ý & BÁO LỖI NGƯỜI DÙNG
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="3. Góp Ý & Báo Lỗi")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3["A1"] = "TỔNG HỢP Ý KIẾN ĐÓNG GÓP, ĐỀ XUẤT TÍNH NĂNG & BÁO LỖI TỪ NGƯỜI DÙNG"
    ws3["A1"].font = title_font
    ws3["A2"] = "Dữ liệu phục vụ nghiên cứu hành vi khách hàng và tối ưu hóa tính năng sản phẩm"
    ws3["A2"].font = subtitle_font
    
    headers3 = [
        "STT", "ID", "Tài Khoản Gửi", "Đánh Giá (⭐)", "Phân Loại Ý Kiến",
        "Nội Dung Phản Hồi / Vấn Đề Gặp Phải", "Thời Gian Gửi", "Trạng Thái Xử Lý", "Ghi Chú Của Admin"
    ]
    for col_idx, h in enumerate(headers3, 1):
        cell = ws3.cell(row=4, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    for row_idx, f in enumerate(feedbacks, 5):
        row_vals3 = [
            row_idx - 4,
            f.get("id", ""),
            f.get("user_id", ""),
            f"{f.get('rating', 5)} Sao",
            f.get("category", ""),
            f.get("comment", ""),
            f.get("created_at", ""),
            f.get("status", "Chờ xử lý"),
            f.get("admin_note", "")
        ]
        for col_idx, val in enumerate(row_vals3, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx in [1, 2, 4, 7, 8]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    # -------------------------------------------------------------
    # SHEET 4: TỔNG QUAN PHÂN TÍCH & ĐỀ XUẤT CẢI THIỆN
    # -------------------------------------------------------------
    ws4 = wb.create_sheet(title="4. Tổng Quan & Đề Xuất")
    ws4.views.sheetView[0].showGridLines = True
    
    ws4["A1"] = "BÁO CÁO PHÂN TÍCH TỔNG QUAN & HƯỚNG CẢI THIỆN SẢN PHẨM"
    ws4["A1"].font = title_font
    ws4["A2"] = "Tổng hợp các chỉ số trọng yếu giúp định hướng nâng cấp hệ thống bóc tách hóa đơn"
    ws4["A2"].font = subtitle_font
    
    ws4["A4"] = "CHỈ SỐ TRỌNG YẾU HỆ THỐNG"
    ws4["B4"] = "GIÁ TRỊ HIỆN TẠI"
    ws4.cell(row=4, column=1).fill = header_fill
    ws4.cell(row=4, column=1).font = header_font
    ws4.cell(row=4, column=2).fill = header_fill
    ws4.cell(row=4, column=2).font = header_font
    
    total_users_cnt = len(users)
    active_users_cnt = len([u for u in users if u.get("days_remaining", 0) > 0])
    total_fbs_cnt = len(feedbacks)
    avg_stars = round(sum(f.get("rating", 5) for f in feedbacks) / total_fbs_cnt, 2) if total_fbs_cnt > 0 else 5.0
    
    kpis = [
        ("Tổng số tài khoản đăng ký", f"{total_users_cnt} tài khoản"),
        ("Tài khoản đang trong hạn 30 ngày dùng thử", f"{active_users_cnt} tài khoản"),
        ("Tổng số ý kiến đóng góp & báo lỗi nhận được", f"{total_fbs_cnt} lượt"),
        ("Mức độ hài lòng trung bình", f"{avg_stars} / 5.0 ⭐"),
        ("Thời gian xuất báo cáo", datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    ]
    
    for r_i, (k, v) in enumerate(kpis, 5):
        c1 = ws4.cell(row=r_i, column=1, value=k)
        c2 = ws4.cell(row=r_i, column=2, value=v)
        c1.font = regular_font
        c2.font = bold_font
        c1.border = thin_border
        c2.border = thin_border
        
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row in [1, 2]:
                    continue
                if cell.value:
                    val_str = str(cell.value)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

if __name__ == "__main__":
    snapshot_path = save_weekly_report_snapshot()
    print(f"Weekly feedback summary generated and saved cleanly at: {snapshot_path}")
