import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any

class InvoiceExporter:
    @staticmethod
    def export_comprehensive_excel(invoices: List[Dict[str, Any]], db, lang: str = "vi") -> bytes:
        """
        Báo cáo Excel Hóa Đơn Chuẩn Kế Toán (TT 91/2026/TT-BTC & NĐ 123/2020/NĐ-CP):
        - Hỗ trợ đa ngôn ngữ: 'vi' (Tiếng Việt), 'en' (English), 'zh' (中文), 'bilingual_zh' (Song ngữ Việt-Trung).
        - Font chữ: Times New Roman đồng nhất toàn bộ bảng.
        - Không dùng gộp ô (Merge & Center) -> Dùng Center Across Selection (centerContinuous).
        - Cố định dòng cột (Freeze Panes) tại ô D5 cho cả 2 Sheet.
        - Độ rộng các cột được căn chỉnh vừa vặn chính xác theo dữ liệu bên trong.
        - Đường kẻ viền phân cách các ô và cột bằng nét mảnh màu xám nhẹ (thin light-gray border #BFBFBF).
        - Cột N (Chữ ký số): Xác định rõ ràng 'Đã ký số' hoặc 'Chưa ký số'.
        - Cột O (Mã CQT): Ghi rõ mã cụ thể, 'Không có mã CQT (HĐ không mã)' hoặc 'Chưa có mã CQT'.
        - Cột Q (Mã tra cứu): Ghi đúng mã số bí mật tra cứu, hoặc 'Chưa có mã tra cứu'.
        - Dòng nghi ngờ / Chưa ký số được tô màu nổi bật cảnh báo.
        """
        wb = openpyxl.Workbook()
        
        # Tiêu đề Sheet 1 theo ngôn ngữ
        headers_ws1_map = {
            "vi": [
                "STT", "Ký hiệu mẫu", "Ký hiệu HĐ", "Số hóa đơn", "Ngày lập",
                "Tên Nhà Cung Cấp (Người bán)", "Mã số thuế NCC", "Địa chỉ Người bán",
                "Tên Đơn vị mua", "Mã số thuế Mua",
                "Tiền chưa thuế (VND)", "Tiền thuế GTGT (VND)", "Tổng thanh toán (VND)",
                "Chữ ký số", "Mã của Cơ quan thuế (MCCQT)", "Website tra cứu", "Mã số tra cứu",
                "Đánh giá hợp lệ (TT 91/2026)", "Ghi chú"
            ],
            "en": [
                "No.", "Form Type", "Serial No.", "Invoice No.", "Issue Date",
                "Supplier / Seller Name", "Supplier Tax ID", "Supplier Address",
                "Buyer / Company Name", "Buyer Tax ID",
                "Pre-Tax Amount (VND)", "VAT Amount (VND)", "Total Payable (VND)",
                "Digital Signature", "Tax Authority Code (MCCQT)", "Lookup URL", "Lookup Secret Code",
                "Compliance Status (TT 91/2026)", "Notes"
            ],
            "zh": [
                "序号", "发票类别", "发票代码", "发票号码", "开票日期",
                "供应商（销方名称）", "供应商税号", "销方地址",
                "购买方单位名称", "购买方税号",
                "不含税金额 (VND)", "增值税税额 (VND)", "价税合计总额 (VND)",
                "电子签名", "税局防伪代码 (MCCQT)", "查验网址", "防伪查验码",
                "合规审核结论 (TT 91/2026)", "备注说明"
            ],
            "bilingual_zh": [
                "STT / 序号", "Ký hiệu mẫu / 类别", "Ký hiệu HĐ / 代码", "Số HĐ / 发票号码", "Ngày lập / 开票日期",
                "Tên Nhà Cung Cấp / 供应商名称", "MST NCC / 供应商税号", "Địa chỉ / 销方地址",
                "Tên Đơn vị mua / 购买方", "MST Mua / 购买方税号",
                "Tiền chưa thuế / 不含税额 (VND)", "Tiền thuế / 增值税额 (VND)", "Tổng thanh toán / 价税合计 (VND)",
                "Chữ ký số / 电子签名", "Mã CQT / 税局代码", "Website / 查验网址", "Mã tra cứu / 查验码",
                "Đánh giá / 合规审核", "Ghi chú / 备注"
            ],
            "ko": [
                "연번", "양식 기호", "계산서 기호", "계산서 번호", "발행일자",
                "공급업체명 (매도자)", "공급자 사업자등록번호", "공급자 주소",
                "구매자 상호 (법인명)", "구매자 사업자등록번호",
                "공급가액 (VND)", "부가가치세 (VND)", "총결제금액 (VND)",
                "전자서명 여부", "세무국 코드 (MCCQT)", "조회 웹사이트", "비밀번호 조회코드",
                "법적 적격성 평가 (TT 91/2026)", "비고"
            ],
            "bilingual_ko": [
                "STT / 연번", "Ký hiệu mẫu / 양식", "Ký hiệu HĐ / 기호", "Số HĐ / 계산서번호", "Ngày lập / 발행일자",
                "Tên NCC / 공급업체명", "MST NCC / 사업자번호", "Địa chỉ / 공급자주소",
                "Tên Đơn vị mua / 구매자", "MST Mua / 구매자사업자번호",
                "Tiền chưa thuế / 공급가액 (VND)", "Tiền thuế / 부가세 (VND)", "Tổng thanh toán / 총금액 (VND)",
                "Chữ ký số / 전자서명", "Mã CQT / 세무국코드", "Website / 조회사이트", "Mã tra cứu / 조회코드",
                "Đánh giá / 적격성평가", "Ghi chú / 비고"
            ]
        }
        headers_ws1 = headers_ws1_map.get(lang, headers_ws1_map["vi"])
        
        main_titles = {
            "vi": ("BẢNG KÊ TỔNG HỢP HÓA ĐƠN ĐIỆN TỬ", "Theo quy chuẩn Thông tư số 91/2026/TT-BTC & Nghị định số 123/2020/NĐ-CP", "Bảng_Kê_Tổng_Hợp"),
            "en": ("COMPREHENSIVE E-INVOICE SUMMARY LEDGER", "Conforming to Circular 91/2026/TT-BTC & Decree 123/2020/ND-CP", "Invoice_Summary"),
            "zh": ("电子发票汇总统计总表", "符合越南财政部第91/2026/TT-BTC号通告与第123/2020/ND-CP号法令合规标准", "发票总表_汇总"),
            "bilingual_zh": ("BẢNG KÊ TỔNG HỢP HÓA ĐƠN / 电子发票汇总表", "Quy chuẩn Thông tư 91/2026/TT-BTC & Nghị định 123 / 越南增值税发票合规标准", "Bang_Ke_Tong_Hop"),
            "ko": ("전자세금계산서 총괄 집계표", "베트남 재무부 통로서 91/2026/TT-BTC 및 시행령 123/2020/ND-CP 표준", "계산서_총괄_집계표"),
            "bilingual_ko": ("BẢNG KÊ TỔNG HỢP HÓA ĐƠN / 전자계산서 총괄표", "Thông tư 91/2026/TT-BTC & Nghị định 123 / 베트남 전자계산서 회계 표준", "Bang_Ke_Tong_Hop")
        }
        title_main, title_sub, sheet1_name = main_titles.get(lang, main_titles["vi"])
        
        # ================= FONTS (Times New Roman) =================
        font_main_title = Font(name="Times New Roman", size=14, bold=True, color="1F4E78")
        font_sub_title = Font(name="Times New Roman", size=10, italic=True, color="595959")
        font_header = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Times New Roman", size=11, bold=True)
        font_regular = Font(name="Times New Roman", size=11)
        font_warning = Font(name="Times New Roman", size=11, bold=True, color="C00000")
        
        # ================= FILLS =================
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_vendor = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        fill_subtotal = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        fill_warning_row = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Vàng nhạt cảnh báo
        fill_valid_status = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Xanh lá nhạt
        fill_invalid_status = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Đỏ cam nhạt
        
        # ================= BORDERS (Nét mảnh màu xám nhẹ #BFBFBF) =================
        thin_gray_border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        total_top_bottom_border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='double', color='000000')
        )

        # =========================================================================
        # SHEET 1: BẢNG KÊ TỔNG HỢP HÓA ĐƠN
        # =========================================================================
        ws1 = wb.active
        ws1.title = sheet1_name
        ws1.views.sheetView[0].showGridLines = True
        
        # Freeze Panes tại D5 (Cố định 3 cột đầu A, B, C và 4 dòng tiêu đề trên)
        ws1.freeze_panes = "D5"
        num_cols1 = len(headers_ws1)

        # Title Row 1 & 2 - DÙNG CENTER ACROSS SELECTION (KHÔNG GỘP Ô)
        ws1.cell(row=1, column=1, value=title_main)
        for c in range(1, num_cols1 + 1):
            cell = ws1.cell(row=1, column=c)
            cell.font = font_main_title
            cell.alignment = Alignment(horizontal="centerContinuous", vertical="center")
            
        ws1.cell(row=2, column=1, value=title_sub)
        for c in range(1, num_cols1 + 1):
            cell = ws1.cell(row=2, column=c)
            cell.font = font_sub_title
            cell.alignment = Alignment(horizontal="centerContinuous", vertical="center")
            
        ws1.row_dimensions[1].height = 24
        ws1.row_dimensions[2].height = 18
        ws1.row_dimensions[4].height = 28
        
        # Header Row 4
        for col_idx, h in enumerate(headers_ws1, 1):
            cell = ws1.cell(row=4, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_gray_border
            
        start_row = 5
        current_row = start_row
        
        for idx, inv in enumerate(invoices, 1):
            is_valid = inv.get("is_valid", 1) == 1
            has_sig = bool(inv.get("has_signature"))
            sig_txt = "Đã ký số" if has_sig else "Chưa ký số"
            status_txt = inv.get("status_summary", "Hợp lệ")
            
            is_suspicious = (not is_valid) or (not has_sig) or ("Nghi ngờ" in status_txt) or ("Không hợp lệ" in status_txt)
            
            row_font = font_warning if is_suspicious else font_regular
            row_fill = fill_warning_row if is_suspicious else None
            
            ws1.cell(row=current_row, column=1, value=idx).alignment = Alignment(horizontal="center")
            ws1.cell(row=current_row, column=2, value=inv.get("kh_mau", "1")).alignment = Alignment(horizontal="center")
            ws1.cell(row=current_row, column=3, value=inv.get("kh_hd", "")).alignment = Alignment(horizontal="center")
            ws1.cell(row=current_row, column=4, value=inv.get("so_hd", "")).alignment = Alignment(horizontal="center")
            ws1.cell(row=current_row, column=5, value=inv.get("ngay_lap", "")).alignment = Alignment(horizontal="center")
            ws1.cell(row=current_row, column=6, value=inv.get("ten_nban", ""))
            ws1.cell(row=current_row, column=7, value=inv.get("mst_nban", "")).alignment = Alignment(horizontal="center")
            ws1.cell(row=current_row, column=8, value=inv.get("dc_nban", ""))
            ws1.cell(row=current_row, column=9, value=inv.get("ten_nmua", "CÔNG TY TNHH TOYO SOLAR"))
            ws1.cell(row=current_row, column=10, value=inv.get("mst_nmua", "2601084657")).alignment = Alignment(horizontal="center")
            
            c_pre = ws1.cell(row=current_row, column=11, value=float(inv.get("tien_chua_thue", 0.0)))
            c_vat = ws1.cell(row=current_row, column=12, value=float(inv.get("tien_thue", 0.0)))
            c_tot = ws1.cell(row=current_row, column=13, value=float(inv.get("tong_tien", 0.0)))
            
            c_pre.number_format = '#,##0'
            c_vat.number_format = '#,##0'
            c_tot.number_format = '#,##0'
            
            # Cột N: Chữ ký số
            c_sig = ws1.cell(row=current_row, column=14, value=sig_txt)
            c_sig.alignment = Alignment(horizontal="center")
            if not has_sig:
                c_sig.font = font_warning
            
            # Cột O: Mã CQT
            ma_cqt_val = inv.get("ma_cqt") or "Chưa có mã CQT"
            ws1.cell(row=current_row, column=15, value=ma_cqt_val).alignment = Alignment(horizontal="center")
            
            # Cột P: Website tra cứu
            ws1.cell(row=current_row, column=16, value=inv.get("web_tra_cuu", "https://hoadondientu.gdt.gov.vn"))
            
            # Cột Q: Mã số tra cứu
            ma_tc_val = inv.get("ma_tra_cuu") or "Chưa có mã tra cứu"
            ws1.cell(row=current_row, column=17, value=ma_tc_val).alignment = Alignment(horizontal="center")
            
            # Cột R: Đánh giá hợp lệ TT 91/2026
            c_status = ws1.cell(row=current_row, column=18, value=status_txt)
            c_status.alignment = Alignment(horizontal="center")
            if is_suspicious:
                c_status.fill = fill_invalid_status
                c_status.font = font_warning
            else:
                c_status.fill = fill_valid_status
                
            # Cột S: Ghi chú
            ws1.cell(row=current_row, column=19, value=inv.get("notes", "Hợp lệ"))
            
            for c in range(1, num_cols1 + 1):
                cell_item = ws1.cell(row=current_row, column=c)
                cell_item.border = thin_gray_border
                if is_suspicious:
                    cell_item.font = font_warning
                    if row_fill and c != 18:
                        cell_item.fill = row_fill
                else:
                    cell_item.font = font_regular
                    
            current_row += 1
            
        # Summary Row (KHÔNG DÙNG MERGE, DÙNG CENTER ACROSS SELECTION CHO CỘT A-J)
        ws1.cell(row=current_row, column=1, value="TỔNG CỘNG")
        for c in range(1, 11):
            cell_lbl = ws1.cell(row=current_row, column=c)
            cell_lbl.font = font_bold
            cell_lbl.alignment = Alignment(horizontal="centerContinuous", vertical="center")
            
        sum_pre = ws1.cell(row=current_row, column=11, value=f"=SUM(K{start_row}:K{current_row-1})")
        sum_vat = ws1.cell(row=current_row, column=12, value=f"=SUM(L{start_row}:L{current_row-1})")
        sum_tot = ws1.cell(row=current_row, column=13, value=f"=SUM(M{start_row}:M{current_row-1})")
        
        for c_sum in [sum_pre, sum_vat, sum_tot]:
            c_sum.number_format = '#,##0'
            c_sum.font = font_bold
            
        for c in range(1, num_cols1 + 1):
            cell_t = ws1.cell(row=current_row, column=c)
            cell_t.border = total_top_bottom_border
            cell_t.fill = fill_subtotal
            
        # ĐẶT ĐỘ RỘNG CỘT CHUẨN XÁC, VỪA VẶN THEO DỮ LIỆU
        exact_widths_ws1 = {
            "A": 8,   # STT
            "B": 14,  # Ký hiệu mẫu
            "C": 14,  # Ký hiệu HĐ
            "D": 16,  # Số hóa đơn
            "E": 14,  # Ngày lập
            "F": 42,  # Tên Nhà Cung Cấp
            "G": 18,  # Mã số thuế NCC
            "H": 38,  # Địa chỉ Người bán
            "I": 32,  # Tên Đơn vị mua
            "J": 16,  # Mã số thuế Mua
            "K": 20,  # Tiền chưa thuế
            "L": 18,  # Tiền thuế GTGT
            "M": 20,  # Tổng thanh toán
            "N": 16,  # Chữ ký số
            "O": 42,  # Mã của Cơ quan thuế
            "P": 38,  # Website tra cứu
            "Q": 24,  # Mã số tra cứu
            "R": 28,  # Đánh giá hợp lệ TT 91/2026
            "S": 38   # Ghi chú
        }
        for col_l, w in exact_widths_ws1.items():
            ws1.column_dimensions[col_l].width = w

        # =========================================================================
        # SHEET 2: CHI TIẾT THEO NHÀ CUNG CẤP
        # =========================================================================
        headers_ws2_map = {
            "vi": [
                "STT", "Nhà Cung Cấp / Tên Hàng Hóa", "Mã Số Thuế", "Số HĐ", "Ký hiệu HĐ", "Ngày lập HĐ",
                "ĐVT", "Số lượng", "Đơn giá (VND)", "Thành tiền (VND)", "Thuế suất", "Tiền thuế GTGT (VND)", "Tổng thanh toán (VND)"
            ],
            "en": [
                "No.", "Supplier / Item Description", "Tax ID", "Inv No.", "Serial", "Inv Date",
                "UOM", "Quantity", "Unit Price (VND)", "Net Amount (VND)", "VAT Rate", "VAT Amount (VND)", "Total Payable (VND)"
            ],
            "zh": [
                "序号", "供应商名称 / 货物及劳务名称", "税号", "发票号码", "发票代码", "开票日期",
                "单位", "数量", "单价 (VND)", "金额 (VND)", "税率", "税额 (VND)", "价税合计 (VND)"
            ],
            "bilingual_zh": [
                "STT / 序号", "Tên Hàng Hóa / 货物劳务名称", "MST / 税号", "Số HĐ / 发票号码", "Ký hiệu / 代码", "Ngày lập / 日期",
                "ĐVT / 单位", "Số lượng / 数量", "Đơn giá / 单价 (VND)", "Thành tiền / 金额 (VND)", "Thuế suất / 税率", "Tiền thuế / 税额 (VND)", "Tổng cộng / 价税合计 (VND)"
            ],
            "ko": [
                "연번", "공급업체명 / 품목 및 용역명", "사업자등록번호", "계산서번호", "기호", "발행일자",
                "단위", "수량", "단가 (VND)", "공급가액 (VND)", "세율", "부가가치세 (VND)", "총결제금액 (VND)"
            ],
            "bilingual_ko": [
                "STT / 연번", "Tên Hàng Hóa / 품목명", "MST / 사업자번호", "Số HĐ / 계산서번호", "Ký hiệu / 기호", "Ngày lập / 발행일",
                "ĐVT / 단위", "Số lượng / 수량", "Đơn giá / 단가 (VND)", "Thành tiền / 공급가액 (VND)", "Thuế suất / 세율", "Tiền thuế / 부가세 (VND)", "Tổng cộng / 총금액 (VND)"
            ]
        }
        headers_ws2 = headers_ws2_map.get(lang, headers_ws2_map["vi"])

        sub_titles_s2 = {
            "vi": ("BẢNG KÊ CHI TIẾT HÀNG HÓA - DỊCH VỤ THEO NHÀ CUNG CẤP", "Phân rã chi tiết từng dòng mặt hàng, đơn giá, thuế suất theo từng hóa đơn và NCC", "Chi_Tiết_Theo_Nhà_Cung_Cấp"),
            "en": ("DETAILED GOODS & SERVICES BREAKDOWN BY SUPPLIER", "Itemized breakdown of unit prices, VAT rates and line totals by supplier", "Supplier_Item_Breakdown"),
            "zh": ("供应商货物与服务采购明细分类总表", "按发票与供应商详细拆分各项商品规格、单价、税率与税额", "供应商采购明细表"),
            "bilingual_zh": ("CHI TIẾT THEO NHÀ CUNG CẤP / 供应商明细表", "Phân rã chi tiết hàng hóa theo NCC / 供应商采购明细分录", "Chi_Tiet_NCC"),
            "ko": ("공급업체별 품목 및 용역 매입 명세서", "계산서 및 공급업체별 단가, 세율, 세액 상세 분할 명세", "공급업체별_매입명세표"),
            "bilingual_ko": ("CHI TIẾT THEO NHÀ CUNG CẤP / 공급업체별 명세", "Phân rã chi tiết hàng hóa theo NCC / 공급업체별 매입명세", "Chi_Tiet_NCC")
        }
        title2_main, title2_sub, sheet2_name = sub_titles_s2.get(lang, sub_titles_s2["vi"])

        ws2 = wb.create_sheet(title=sheet2_name)
        ws2.views.sheetView[0].showGridLines = True
        
        # Freeze Panes tại D5
        ws2.freeze_panes = "D5"
        num_cols2 = len(headers_ws2)
        
        # Title Row 1 & 2 (Center Continuous)
        ws2.cell(row=1, column=1, value=title2_main)
        for c in range(1, num_cols2 + 1):
            cell = ws2.cell(row=1, column=c)
            cell.font = font_main_title
            cell.alignment = Alignment(horizontal="centerContinuous", vertical="center")
            
        ws2.cell(row=2, column=1, value=title2_sub)
        for c in range(1, num_cols2 + 1):
            cell = ws2.cell(row=2, column=c)
            cell.font = font_sub_title
            cell.alignment = Alignment(horizontal="centerContinuous", vertical="center")
            
        ws2.row_dimensions[1].height = 24
        ws2.row_dimensions[2].height = 18
        ws2.row_dimensions[4].height = 28
        
        for col_idx, h in enumerate(headers_ws2, 1):
            cell = ws2.cell(row=4, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_gray_border
            
        supplier_dict = {}
        for inv in invoices:
            key = (inv.get("ten_nban", "Chưa xác định"), inv.get("mst_nban", ""))
            if key not in supplier_dict:
                supplier_dict[key] = []
            supplier_dict[key].append(inv)
            
        r2 = 5
        global_item_idx = 1
        
        for (sup_name, sup_mst), sup_invoices in supplier_dict.items():
            # Supplier Group Header (KHÔNG DÙNG MERGE, DÙNG CENTER ACROSS SELECTION)
            ws2.cell(row=r2, column=1, value="🏢 NCC:")
            ws2.cell(row=r2, column=2, value=f"{sup_name} (MST: {sup_mst})")
            for c in range(1, num_cols2 + 1):
                cell_v = ws2.cell(row=r2, column=c)
                cell_v.font = font_bold
                cell_v.fill = fill_vendor
                cell_v.border = thin_gray_border
                if c >= 2:
                    cell_v.alignment = Alignment(horizontal="left", vertical="center")
            r2 += 1
            
            sup_start_item_row = r2
            
            for inv in sup_invoices:
                inv_id = inv.get("id")
                items = db.get_invoice_items(inv_id) if inv_id else inv.get("items", [])
                
                if not items:
                    items = [{
                        "stt": 1,
                        "ten_hang": f"Dịch vụ / Hàng hóa theo HĐ {inv.get('so_hd')}",
                        "dvt": "Gói",
                        "so_luong": 1.0,
                        "don_gia": float(inv.get("tien_chua_thue", 0.0)),
                        "thanh_tien": float(inv.get("tien_chua_thue", 0.0)),
                        "thue_suat": "0%" if float(inv.get("tien_thue", 0.0)) == 0 else "10%",
                        "tien_thue": float(inv.get("tien_thue", 0.0))
                    }]
                    
                for it in items:
                    ws2.cell(row=r2, column=1, value=global_item_idx).alignment = Alignment(horizontal="center")
                    ws2.cell(row=r2, column=2, value=it.get("ten_hang", ""))
                    ws2.cell(row=r2, column=3, value=sup_mst).alignment = Alignment(horizontal="center")
                    ws2.cell(row=r2, column=4, value=inv.get("so_hd", "")).alignment = Alignment(horizontal="center")
                    ws2.cell(row=r2, column=5, value=inv.get("kh_hd", "")).alignment = Alignment(horizontal="center")
                    ws2.cell(row=r2, column=6, value=inv.get("ngay_lap", "")).alignment = Alignment(horizontal="center")
                    ws2.cell(row=r2, column=7, value=it.get("dvt", "")).alignment = Alignment(horizontal="center")
                    
                    c_sl = ws2.cell(row=r2, column=8, value=float(it.get("so_luong", 1.0)))
                    c_dg = ws2.cell(row=r2, column=9, value=float(it.get("don_gia", 0.0)))
                    c_tt = ws2.cell(row=r2, column=10, value=float(it.get("thanh_tien", 0.0)))
                    
                    ws2.cell(row=r2, column=11, value=str(it.get("thue_suat", "0%"))).alignment = Alignment(horizontal="center")
                    c_tv = ws2.cell(row=r2, column=12, value=float(it.get("tien_thue", 0.0)))
                    c_all = ws2.cell(row=r2, column=13, value=float(it.get("thanh_tien", 0.0)) + float(it.get("tien_thue", 0.0)))
                    
                    c_sl.number_format = '#,##0.##'
                    c_dg.number_format = '#,##0'
                    c_tt.number_format = '#,##0'
                    c_tv.number_format = '#,##0'
                    c_all.number_format = '#,##0'
                    
                    for c in range(1, num_cols2 + 1):
                        cell_item = ws2.cell(row=r2, column=c)
                        cell_item.border = thin_gray_border
                        cell_item.font = font_regular
                    r2 += 1
                    global_item_idx += 1
                    
            # Subtotal per Supplier (Center Continuous across A-I)
            ws2.cell(row=r2, column=2, value=f"Cộng Nhà Cung Cấp [{sup_name}]:")
            for c in range(1, 10):
                cell_sub_lbl = ws2.cell(row=r2, column=c)
                cell_sub_lbl.font = font_bold
                if c >= 2:
                    cell_sub_lbl.alignment = Alignment(horizontal="right", vertical="center")
                    
            sub_tt = ws2.cell(row=r2, column=10, value=f"=SUM(J{sup_start_item_row}:J{r2-1})")
            sub_tv = ws2.cell(row=r2, column=12, value=f"=SUM(L{sup_start_item_row}:L{r2-1})")
            sub_all = ws2.cell(row=r2, column=13, value=f"=SUM(M{sup_start_item_row}:M{r2-1})")
            
            for c_sub in [sub_tt, sub_tv, sub_all]:
                c_sub.number_format = '#,##0'
                c_sub.font = font_bold
                
            for c in range(1, num_cols2 + 1):
                cell_sub = ws2.cell(row=r2, column=c)
                cell_sub.fill = fill_subtotal
                cell_sub.border = thin_gray_border
            r2 += 1
            
        exact_widths_ws2 = {
            "A": 8,   # STT
            "B": 48,  # NCC / Tên Hàng
            "C": 18,  # Mã Số Thuế
            "D": 16,  # Số HĐ
            "E": 15,  # Ký hiệu HĐ
            "F": 14,  # Ngày lập HĐ
            "G": 10,  # ĐVT
            "H": 12,  # Số lượng
            "I": 18,  # Đơn giá
            "J": 20,  # Thành tiền
            "K": 12,  # Thuế suất
            "L": 18,  # Tiền thuế GTGT
            "M": 20   # Tổng thanh toán
        }
        for col_l, w in exact_widths_ws2.items():
            ws2.column_dimensions[col_l].width = w
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
